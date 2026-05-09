from typing import Optional
from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import NoticeVO, NoticePageParam, NoticeExportParam, NoticeImportParam
from ...service import NoticeService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/notice/page",
    summary="获取通知分页",
    response_model=Result[PageData[NoticeVO]]
)
@HeiCheckPermission("sys:notice:page")
async def page(
    request: Request,
    current: int = Query(default=1),
    size: int = Query(default=10),
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    return success(service.page(NoticePageParam(current=current, size=size)))


@router.post(
    "/api/v1/sys/notice/create",
    summary="添加通知",
    response_model=Result
)
@HeiCheckPermission("sys:notice:create")
async def create(
    request: Request,
    vo: NoticeVO,
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/notice/modify",
    summary="编辑通知",
    response_model=Result
)
@HeiCheckPermission("sys:notice:modify")
async def modify(
    request: Request,
    vo: NoticeVO,
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/notice/remove",
    summary="删除通知",
    response_model=Result
)
@HeiCheckPermission("sys:notice:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/notice/detail",
    summary="获取通知详情",
    response_model=Result[NoticeVO]
)
@HeiCheckPermission("sys:notice:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/notice/export",
    summary="导出通知数据")
@HeiCheckPermission("sys:notice:export")
async def export(
    request: Request,
    export_type: str = Query(default="current"),
    current: Optional[int] = Query(default=None),
    size: Optional[int] = Query(default=None),
    selected_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    param = NoticeExportParam(
        export_type=export_type,
        current=current,
        size=size,
        selected_id=selected_id.split(",") if selected_id else None
    )
    return service.export(param)


@router.get(
    "/api/v1/sys/notice/template",
    summary="下载通知导入模板")
@HeiCheckPermission("sys:notice:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = NoticeService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/notice/import",
    summary="导入通知数据",
    response_model=Result
)
@HeiCheckPermission("sys:notice:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(NoticeVO(**row_dict))

    service = NoticeService(db)
    result = await service.import_data(NoticeImportParam(data=data_list), request)
    return success(result)
