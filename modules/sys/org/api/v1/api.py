from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import OrgVO, OrgPageParam, OrgTreeParam, OrgExportParam, OrgImportParam, GrantOrgRoleParam
from ...service import OrgService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/org/page",
    summary="获取组织分页",
    response_model=Result[PageData[OrgVO]]
)
@HeiCheckPermission("sys:org:page")
async def page(
    request: Request,
    param: OrgPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    return success(service.page(param))


@router.get(
    "/api/v1/sys/org/tree",
    summary="获取组织树",
    response_model=Result[list[dict]]
)
@HeiCheckPermission("sys:org:tree")
async def tree(
    request: Request,
    param: OrgTreeParam = Depends(),
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    return success(service.tree(param))


@router.post(
    "/api/v1/sys/org/create",
    summary="添加组织",
    response_model=Result
)
@HeiCheckPermission("sys:org:create")
async def create(
    request: Request,
    vo: OrgVO,
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/org/modify",
    summary="编辑组织",
    response_model=Result
)
@HeiCheckPermission("sys:org:modify")
async def modify(
    request: Request,
    vo: OrgVO,
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/org/remove",
    summary="删除组织",
    response_model=Result
)
@HeiCheckPermission("sys:org:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/org/detail",
    summary="获取组织详情",
    response_model=Result[OrgVO]
)
@HeiCheckPermission("sys:org:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/org/export",
    summary="导出组织数据")
@HeiCheckPermission("sys:org:export")
async def export(
    request: Request,
    param: OrgExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/org/template",
    summary="下载组织导入模板")
@HeiCheckPermission("sys:org:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/org/import",
    summary="导入组织数据",
    response_model=Result
)
@HeiCheckPermission("sys:org:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(OrgVO(**row_dict))

    service = OrgService(db)
    result = await service.import_data(OrgImportParam(data=data_list), request)
    return success(result)


@router.post(
    "/api/v1/sys/org/grant-role",
    summary="分配组织角色",
    response_model=Result
)
@HeiCheckPermission("sys:org:grant-role")
async def grant_role(
    request: Request,
    param: GrantOrgRoleParam,
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    await service.grant_roles(param, request)
    return success()


@router.get(
    "/api/v1/sys/org/own-roles",
    summary="获取组织已分配的角色ID列表"
)
@HeiCheckPermission("sys:org:own-roles")
async def own_roles(
    request: Request,
    org_id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = OrgService(db)
    return success(service.get_org_role_ids(org_id))
