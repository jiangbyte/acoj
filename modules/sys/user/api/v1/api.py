from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission, HeiCheckLogin
from core.utils.excel_utils import validate_import_file
from ...params import UserVO, UserPageParam, UserExportParam, UserImportParam, GrantRoleParam, GrantGroupParam
from ...service import UserService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/user/page",
    summary="获取用户分页",
    response_model=Result[PageData[UserVO]]
)
@HeiCheckPermission("sys:user:page")
async def page(
    request: Request,
    param: UserPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = UserService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/user/create",
    summary="添加用户",
    response_model=Result
)
@HeiCheckPermission("sys:user:create")
async def create(
    request: Request,
    vo: UserVO,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/user/modify",
    summary="编辑用户",
    response_model=Result
)
@HeiCheckPermission("sys:user:modify")
async def modify(
    request: Request,
    vo: UserVO,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/user/remove",
    summary="删除用户",
    response_model=Result
)
@HeiCheckPermission("sys:user:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/user/detail",
    summary="获取用户详情",
    response_model=Result[UserVO]
)
@HeiCheckPermission("sys:user:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = UserService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/user/export",
    summary="导出用户数据")
@HeiCheckPermission("sys:user:export")
async def export(
    request: Request,
    param: UserExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = UserService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/user/template",
    summary="下载用户导入模板")
@HeiCheckPermission("sys:user:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/user/import",
    summary="导入用户数据",
    response_model=Result
)
@HeiCheckPermission("sys:user:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(UserVO(**row_dict))

    service = UserService(db)
    result = await service.import_data(UserImportParam(data=data_list), request)
    return success(result)


@router.post(
    "/api/v1/sys/user/grant-role",
    summary="分配用户角色",
    response_model=Result
)
@HeiCheckPermission("sys:user:grant-role")
async def grant_role(
    request: Request,
    param: GrantRoleParam,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    await service.grant_roles(param, request)
    return success()


@router.post(
    "/api/v1/sys/user/grant-group",
    summary="分配用户用户组",
    response_model=Result
)
@HeiCheckPermission("sys:user:grant-group")
async def grant_group(
    request: Request,
    param: GrantGroupParam,
    db: Session = Depends(get_db)
):
    service = UserService(db)
    await service.grant_groups(param, request)
    return success()




@router.get(
    "/api/v1/sys/user/own-roles",
    summary="获取用户已分配的角色ID列表"
)
@HeiCheckPermission("sys:user:own-roles")
async def own_roles(
    request: Request,
    user_id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = UserService(db)
    return success(service.get_user_role_ids(user_id))


@router.get(
    "/api/v1/sys/user/own-groups",
    summary="获取用户已分配的用户组ID列表"
)
@HeiCheckPermission("sys:user:own-groups")
async def own_groups(
    request: Request,
    user_id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = UserService(db)
    return success(service.get_user_group_ids(user_id))


@router.get(
    "/api/v1/sys/user/current",
    summary="获取当前用户信息",
)
@HeiCheckLogin
async def get_current_user(request: Request, db: Session = Depends(get_db)):
    service = UserService(db)
    data = await service.get_current_user(request)
    return success(data)


@router.get(
    "/api/v1/sys/user/menus",
    summary="获取当前用户菜单树",
)
@HeiCheckLogin
async def get_current_user_menus(request: Request, db: Session = Depends(get_db)):
    service = UserService(db)
    data = await service.get_current_user_menus(request)
    return success(data)


@router.get(
    "/api/v1/sys/user/permissions",
    summary="获取当前用户权限码列表",
)
@HeiCheckLogin
async def get_current_user_permissions(request: Request, db: Session = Depends(get_db)):
    service = UserService(db)
    data = await service.get_current_user_permissions(request)
    return success(data)


