from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import GroupVO, GroupPageParam, GroupTreeParam, GroupExportParam, GroupImportParam
from ...service import GroupService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/group/page",
    summary="获取用户组分页",
    response_model=Result[PageData[GroupVO]]
)
@HeiCheckPermission("sys:group:page")
async def page(
    request: Request,
    param: GroupPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    return success(service.page(param))


@router.get(
    "/api/v1/sys/group/union-tree",
    summary="获取组织与用户组联合树",
    response_model=Result[list[dict]]
)
@HeiCheckPermission("sys:group:tree")
async def union_tree(
    request: Request,
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    return success(service.union_tree())


@router.get(
    "/api/v1/sys/group/tree",
    summary="获取用户组树",
    response_model=Result[list[dict]]
)
@HeiCheckPermission("sys:group:tree")
async def tree(
    request: Request,
    param: GroupTreeParam = Depends(),
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    return success(service.tree(param))


@router.post(
    "/api/v1/sys/group/create",
    summary="添加用户组",
    response_model=Result
)
@HeiCheckPermission("sys:group:create")
async def create(
    request: Request,
    vo: GroupVO,
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/group/modify",
    summary="编辑用户组",
    response_model=Result
)
@HeiCheckPermission("sys:group:modify")
async def modify(
    request: Request,
    vo: GroupVO,
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/group/remove",
    summary="删除用户组",
    response_model=Result
)
@HeiCheckPermission("sys:group:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/group/detail",
    summary="获取用户组详情",
    response_model=Result[GroupVO]
)
@HeiCheckPermission("sys:group:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/group/export",
    summary="导出用户组数据")
@HeiCheckPermission("sys:group:export")
async def export(
    request: Request,
    param: GroupExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/group/template",
    summary="下载用户组导入模板")
@HeiCheckPermission("sys:group:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = GroupService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/group/import",
    summary="导入用户组数据",
    response_model=Result
)
@HeiCheckPermission("sys:group:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(GroupVO(**row_dict))

    service = GroupService(db)
    result = await service.import_data(GroupImportParam(data=data_list), request)
    return success(result)


# 用户组-角色关联已废弃（使用 ral_role_permission.scope 的 GROUP / CUSTOM_GROUP）
