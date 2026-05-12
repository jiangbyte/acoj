from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import PermissionVO, PermissionPageParam, PermissionExportParam, PermissionImportParam
from ...service import PermissionService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/permission/page",
    summary="获取权限分页",
    response_model=Result[PageData[PermissionVO]]
)
@HeiCheckPermission("sys:permission:page")
async def page(
    request: Request,
    param: PermissionPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/permission/create",
    summary="添加权限",
    response_model=Result
)
@HeiCheckPermission("sys:permission:create")
async def create(
    request: Request,
    vo: PermissionVO,
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/permission/modify",
    summary="编辑权限",
    response_model=Result
)
@HeiCheckPermission("sys:permission:modify")
async def modify(
    request: Request,
    vo: PermissionVO,
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/permission/remove",
    summary="删除权限",
    response_model=Result
)
@HeiCheckPermission("sys:permission:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/permission/detail",
    summary="获取权限详情",
    response_model=Result[PermissionVO]
)
@HeiCheckPermission("sys:permission:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/permission/export",
    summary="导出权限数据")
@HeiCheckPermission("sys:permission:export")
async def export(
    request: Request,
    param: PermissionExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/permission/template",
    summary="下载权限导入模板")
@HeiCheckPermission("sys:permission:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/permission/import",
    summary="导入权限数据",
    response_model=Result
)
@HeiCheckPermission("sys:permission:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(PermissionVO(**row_dict))
    service = PermissionService(db)
    result = await service.import_data(PermissionImportParam(data=data_list), request)
    return success(result)


@router.get(
    "/api/v1/sys/permission/modules",
    summary="获取权限模块列表")
@HeiCheckPermission("sys:permission:modules")
async def list_modules(
    request: Request,
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    return success(await service.list_modules())


@router.get(
    "/api/v1/sys/permission/by-module",
    summary="根据模块获取权限列表")
@HeiCheckPermission("sys:permission:by-module")
async def by_module(
    request: Request,
    module: str = Query(...),
    db: Session = Depends(get_db)
):
    service = PermissionService(db)
    return success(await service.list_permissions_by_module(module))


@router.get(
    "/api/v1/sys/permission/list",
    summary="获取全部权限列表（用于选择器）"
)
@HeiCheckPermission("sys:permission:page")
async def list_all(
    request: Request,
    db: Session = Depends(get_db)
):
    return success(PermissionService(db).list_all())
