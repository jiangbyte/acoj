from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import PositionVO, PositionPageParam, PositionExportParam, PositionImportParam
from ...service import PositionService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/position/page",
    summary="获取职位分页",
    response_model=Result[PageData[PositionVO]]
)
@HeiCheckPermission("sys:position:page")
async def page(
    request: Request,
    param: PositionPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/position/create",
    summary="添加职位",
    response_model=Result
)
@HeiCheckPermission("sys:position:create")
async def create(
    request: Request,
    vo: PositionVO,
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/position/modify",
    summary="编辑职位",
    response_model=Result
)
@HeiCheckPermission("sys:position:modify")
async def modify(
    request: Request,
    vo: PositionVO,
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/position/remove",
    summary="删除职位",
    response_model=Result
)
@HeiCheckPermission("sys:position:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/position/detail",
    summary="获取职位详情",
    response_model=Result[PositionVO]
)
@HeiCheckPermission("sys:position:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/position/export",
    summary="导出职位数据")
@HeiCheckPermission("sys:position:export")
async def export(
    request: Request,
    param: PositionExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/position/template",
    summary="下载职位导入模板")
@HeiCheckPermission("sys:position:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = PositionService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/position/import",
    summary="导入职位数据",
    response_model=Result
)
@HeiCheckPermission("sys:position:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(PositionVO(**row_dict))

    service = PositionService(db)
    result = await service.import_data(PositionImportParam(data=data_list), request)
    return success(result)


