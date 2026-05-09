from typing import Optional
from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import DictVO, DictPageParam, DictListParam, DictTreeParam, DictExportParam, DictImportParam
from ...service import DictService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/dict/page",
    summary="获取字典分页",
    response_model=Result[PageData[DictVO]]
)
@HeiCheckPermission("sys:dict:page")
async def page(
    request: Request,
    current: int = Query(default=1),
    size: int = Query(default=10),
    parent_id: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    keyword: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    param = DictPageParam(
        current=current, size=size,
        parent_id=parent_id, category=category, keyword=keyword
    )
    return success(service.page(param))


@router.get(
    "/api/v1/sys/dict/list",
    summary="获取字典列表",
    response_model=Result[list[DictVO]]
)
@HeiCheckPermission("sys:dict:list")
async def dict_list(
    request: Request,
    parent_id: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    param = DictListParam(parent_id=parent_id, category=category)
    return success(service.list(param))


@router.get(
    "/api/v1/sys/dict/tree",
    summary="获取字典树",
    response_model=Result[list[dict]]
)
@HeiCheckPermission("sys:dict:tree")
async def tree(
    request: Request,
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    param = DictTreeParam(category=category)
    return success(service.tree(param))


@router.post(
    "/api/v1/sys/dict/create",
    summary="添加字典",
    response_model=Result
)
@HeiCheckPermission("sys:dict:create")
async def create(
    request: Request,
    vo: DictVO,
    db: Session = Depends(get_db)
):
    service = DictService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/dict/modify",
    summary="编辑字典",
    response_model=Result
)
@HeiCheckPermission("sys:dict:modify")
async def modify(
    request: Request,
    vo: DictVO,
    db: Session = Depends(get_db)
):
    service = DictService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/dict/remove",
    summary="删除字典",
    response_model=Result
)
@HeiCheckPermission("sys:dict:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = DictService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/dict/detail",
    summary="获取字典详情",
    response_model=Result[DictVO]
)
@HeiCheckPermission("sys:dict:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/dict/get-label",
    summary="根据字典编码和值获取字典标签"
)
@HeiCheckPermission("sys:dict:get-label")
async def get_dict_label(
    request: Request,
    type_code: str = Query(...),
    value: str = Query(...),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    label = service.get_dict_label(type_code, value)
    return success({"type_code": type_code, "value": value, "label": label})


@router.get(
    "/api/v1/sys/dict/get-children",
    summary="根据字典编码获取子字典列表"
)
@HeiCheckPermission("sys:dict:get-children")
async def get_dict_children(
    request: Request,
    type_code: str = Query(...),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    return success(service.get_dict_children(type_code))


@router.get(
    "/api/v1/sys/dict/export",
    summary="导出字典数据")
@HeiCheckPermission("sys:dict:export")
async def export(
    request: Request,
    export_type: str = Query(default="current"),
    current: Optional[int] = Query(default=None),
    size: Optional[int] = Query(default=None),
    selected_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    service = DictService(db)
    param = DictExportParam(
        export_type=export_type,
        current=current,
        size=size,
        selected_id=selected_id.split(",") if selected_id else None
    )
    return service.export(param)


@router.get(
    "/api/v1/sys/dict/template",
    summary="下载字典导入模板")
@HeiCheckPermission("sys:dict:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = DictService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/dict/import",
    summary="导入字典数据",
    response_model=Result
)
@HeiCheckPermission("sys:dict:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(DictVO(**row_dict))

    service = DictService(db)
    result = await service.import_data(DictImportParam(data=data_list), request)
    return success(result)
