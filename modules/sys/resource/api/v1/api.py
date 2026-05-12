from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import ModuleVO, ResourceVO, ModulePageParam, ResourcePageParam
from ...params import ModuleExportParam, ResourceExportParam, ModuleImportParam, ResourceImportParam, BindPermissionParam
from ...service import ModuleService, ResourceService
from openpyxl import load_workbook
import io

router = APIRouter()


# ---- Module APIs ----

@router.get(
    "/api/v1/sys/module/page",
    summary="获取模块分页",
    response_model=Result[PageData[ModuleVO]]
)
@HeiCheckPermission("sys:module:page")
async def module_page(
    request: Request,
    param: ModulePageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/module/create",
    summary="添加模块",
    response_model=Result
)
@HeiCheckPermission("sys:module:create")
async def module_create(
    request: Request,
    vo: ModuleVO,
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/module/modify",
    summary="编辑模块",
    response_model=Result
)
@HeiCheckPermission("sys:module:modify")
async def module_modify(
    request: Request,
    vo: ModuleVO,
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/module/remove",
    summary="删除模块",
    response_model=Result
)
@HeiCheckPermission("sys:module:remove")
async def module_remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/module/detail",
    summary="获取模块详情",
    response_model=Result[ModuleVO]
)
@HeiCheckPermission("sys:module:detail")
async def module_detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/module/export",
    summary="导出模块数据")
@HeiCheckPermission("sys:module:export")
async def module_export(
    request: Request,
    param: ModuleExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/module/template",
    summary="下载模块导入模板")
@HeiCheckPermission("sys:module:template")
async def module_download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = ModuleService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/module/import",
    summary="导入模块数据",
    response_model=Result
)
@HeiCheckPermission("sys:module:import")
async def module_import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(ModuleVO(**row_dict))

    service = ModuleService(db)
    result = await service.import_data(ModuleImportParam(data=data_list), request)
    return success(result)


# ---- Resource APIs ----

@router.get(
    "/api/v1/sys/resource/tree",
    summary="获取资源树",
    response_model=Result[list]
)
@HeiCheckPermission("sys:resource:tree")
async def resource_tree(
    request: Request,
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    return success(service.tree())


@router.get(
    "/api/v1/sys/resource/page",
    summary="获取资源分页",
    response_model=Result[PageData[ResourceVO]]
)
@HeiCheckPermission("sys:resource:page")
async def resource_page(
    request: Request,
    param: ResourcePageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/resource/create",
    summary="添加资源",
    response_model=Result
)
@HeiCheckPermission("sys:resource:create")
async def resource_create(
    request: Request,
    vo: ResourceVO,
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/resource/modify",
    summary="编辑资源",
    response_model=Result
)
@HeiCheckPermission("sys:resource:modify")
async def resource_modify(
    request: Request,
    vo: ResourceVO,
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/resource/remove",
    summary="删除资源",
    response_model=Result
)
@HeiCheckPermission("sys:resource:remove")
async def resource_remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/resource/detail",
    summary="获取资源详情",
    response_model=Result[ResourceVO]
)
@HeiCheckPermission("sys:resource:detail")
async def resource_detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/resource/export",
    summary="导出资源数据")
@HeiCheckPermission("sys:resource:export")
async def resource_export(
    request: Request,
    param: ResourceExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/resource/template",
    summary="下载资源导入模板")
@HeiCheckPermission("sys:resource:template")
async def resource_download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = ResourceService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/resource/import",
    summary="导入资源数据",
    response_model=Result
)
@HeiCheckPermission("sys:resource:import")
async def resource_import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(ResourceVO(**row_dict))

    service = ResourceService(db)
    result = await service.import_data(ResourceImportParam(data=data_list), request)
    return success(result)


@router.get(
    "/api/v1/sys/resource/own-permissions",
    summary="获取资源已绑定的权限ID列表"
)
@HeiCheckPermission("sys:resource:detail")
async def resource_own_permissions(
    request: Request,
    resource_id: str = Query(...),
    db: Session = Depends(get_db)
):
    return success(ResourceService(db).get_permission_ids(resource_id))


@router.post(
    "/api/v1/sys/resource/bind-permissions",
    summary="绑定权限到资源（覆盖式替换）",
    response_model=Result
)
@HeiCheckPermission("sys:resource:modify")
async def resource_bind_permissions(
    request: Request,
    param: BindPermissionParam,
    db: Session = Depends(get_db)
):
    await ResourceService(db).bind_permissions(
        param.resource_id, param.permission_ids, request
    )
    return success()
