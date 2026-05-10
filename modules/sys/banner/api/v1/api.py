from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import BannerVO, BannerPageParam, BannerExportParam, BannerImportParam
from ...service import BannerService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/sys/banner/page",
    summary="获取Banner分页",
    response_model=Result[PageData[BannerVO]]
)
@HeiCheckPermission("sys:banner:page")
async def page(
    request: Request,
    param: BannerPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/sys/banner/create",
    summary="添加Banner",
    response_model=Result
)
@HeiCheckPermission("sys:banner:create")
async def create(
    request: Request,
    vo: BannerVO,
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/sys/banner/modify",
    summary="编辑Banner",
    response_model=Result
)
@HeiCheckPermission("sys:banner:modify")
async def modify(
    request: Request,
    vo: BannerVO,
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/sys/banner/remove",
    summary="删除Banner",
    response_model=Result
)
@HeiCheckPermission("sys:banner:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/sys/banner/detail",
    summary="获取Banner详情",
    response_model=Result[BannerVO]
)
@HeiCheckPermission("sys:banner:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/sys/banner/export",
    summary="导出Banner数据")
@HeiCheckPermission("sys:banner:export")
async def export(
    request: Request,
    param: BannerExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    return service.export(param)


@router.get(
    "/api/v1/sys/banner/template",
    summary="下载Banner导入模板")
@HeiCheckPermission("sys:banner:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = BannerService(db)
    return service.download_template()


@router.post(
    "/api/v1/sys/banner/import",
    summary="导入Banner数据",
    response_model=Result
)
@HeiCheckPermission("sys:banner:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(BannerVO(**row_dict))
    
    service = BannerService(db)
    result = await service.import_data(BannerImportParam(data=data_list), request)
    return success(result)
