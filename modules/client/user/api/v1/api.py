from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.orm import Session
from core.result import Result, PageData, success
from core.pojo import IdParam, IdsParam
from core.db import get_db
from core.auth.decorator import HeiClientCheckPermission
from core.utils.excel_utils import validate_import_file
from ...params import ClientUserVO, ClientUserPageParam, ClientUserExportParam, ClientUserImportParam
from ...service import ClientUserService
from openpyxl import load_workbook
import io

router = APIRouter()


@router.get(
    "/api/v1/c/client-user/page",
    summary="获取C端用户分页",
    response_model=Result[PageData[ClientUserVO]]
)
@HeiClientCheckPermission("c:client-user:page")
async def page(
    request: Request,
    param: ClientUserPageParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    return success(service.page(param))


@router.post(
    "/api/v1/c/client-user/create",
    summary="添加C端用户",
    response_model=Result
)
@HeiClientCheckPermission("c:client-user:create")
async def create(
    request: Request,
    vo: ClientUserVO,
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    await service.create(vo, request)
    return success()


@router.post(
    "/api/v1/c/client-user/modify",
    summary="编辑C端用户",
    response_model=Result
)
@HeiClientCheckPermission("c:client-user:modify")
async def modify(
    request: Request,
    vo: ClientUserVO,
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    await service.modify(vo, request)
    return success()


@router.post(
    "/api/v1/c/client-user/remove",
    summary="删除C端用户",
    response_model=Result
)
@HeiClientCheckPermission("c:client-user:remove")
async def remove(
    request: Request,
    param: IdsParam,
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    service.remove(param)
    return success()


@router.get(
    "/api/v1/c/client-user/detail",
    summary="获取C端用户详情",
    response_model=Result[ClientUserVO]
)
@HeiClientCheckPermission("c:client-user:detail")
async def detail(
    request: Request,
    id: str = Query(...),
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    data = service.detail(IdParam(id=id))
    return success(data.model_dump() if data else None)


@router.get(
    "/api/v1/c/client-user/export",
    summary="导出C端用户数据")
@HeiClientCheckPermission("c:client-user:export")
async def export(
    request: Request,
    param: ClientUserExportParam = Depends(),
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    return service.export(param)


@router.get(
    "/api/v1/c/client-user/template",
    summary="下载C端用户导入模板")
@HeiClientCheckPermission("c:client-user:template")
async def download_template(
    request: Request,
    db: Session = Depends(get_db)
):
    service = ClientUserService(db)
    return service.download_template()


@router.post(
    "/api/v1/c/client-user/import",
    summary="导入C端用户数据",
    response_model=Result
)
@HeiClientCheckPermission("c:client-user:import")
async def import_data(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(ClientUserVO(**row_dict))

    service = ClientUserService(db)
    result = await service.import_data(ClientUserImportParam(data=data_list), request)
    return success(result)
