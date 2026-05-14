import io
from typing import List, Type
from urllib.parse import quote
from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from config.settings import settings
from core.exception import BusinessException
from core.result import success
_import_max_bytes = settings.app.import_max_file_size_mb * 1024 * 1024


def validate_import_file(file: UploadFile):
    """Validate uploaded import file size. Raises BusinessException if too large."""
    if file.size and file.size > _import_max_bytes:
        raise BusinessException(f"导入文件大小不能超过{settings.app.import_max_file_size_mb}MB")


def export_excel(data: List[dict], filename: str, sheet_name: str = "Sheet1") -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    encoded_filename = quote(f"{filename}.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


async def handle_import(
    file: UploadFile,
    service_class: Type,
    vo_class: Type[BaseModel],
    import_param_class: Type[BaseModel],
    db: Session,
    request=None,
):
    """Shared import handler for Excel-based data import.

    Parses the uploaded Excel file, converts rows to VO instances,
    and delegates to the service's import_data method.

    Usage in router::

        @router.post(...)
        async def import_data(
            file: UploadFile = File(...),
            db: Session = Depends(get_db),
        ):
            return await handle_import(file, XxxService, XxxVO, XxxImportParam, db)
    """
    validate_import_file(file)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data_list.append(vo_class(**row_dict))

    service = service_class(db)
    result = await service.import_data(import_param_class(data=data_list), request)
    return success(result)
